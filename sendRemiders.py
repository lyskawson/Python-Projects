#! /usr/bin/env python3
import openpyxl, smtplib



password = 'Qahstqw535T?'
wb = openpyxl.load_workbook('/Users/aleklyskawa/Desktop/examples/duesRecords.xlsx')
sheet = wb.active 
maxCol = sheet.max_column
latestMonth = sheet.cell(row=1, column=maxCol).value

unpaidMembers = {}

for r in range(2, sheet.max_row + 1):
    payment = sheet.cell(row=r, column=maxCol).value
    if payment != sheet.cell(row=r, column=maxCol):
        name = sheet.cell(row=r, column=1).value 
        email = sheet.cell(row=r, column=2)
        unpaidMembers[name] = email

smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login('m237238r@gmail.com', password)

for name, email in unpaidMembers.items():
    body = "'Subject: %s dues unpaid.\nDear %s, \nRecords show that you have not paid dues for %s. Please make this payment as soon as possible. Thank you!'" % (latestMonth, name, latestMonth)
    print('Sending email to %s...' % email)
    sendmailStatus = smtpObj.sendmail('m237238r@gmail.com', email, body)
    
   
    if sendmailStatus != {}:
        print('There was a problem sending email to %s: %s' % (email, sendmailStatus))

    
smtpObj.quit()





